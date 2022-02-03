from openpyxl import load_workbook

class File:
    def __init__(self, filename, path = '../input', sheet_name='Arkusz1'):
        '''this is file class'''
        '''file variables'''
        self.filename = filename
        self.path = path
        self.sheet_name = sheet_name
        '''on start actions'''
        self.open_file()

    def open_file(self):
        '''dhis function open file (work sheet) as ws'''
        self.wb = load_workbook(filename=self.path + '/' + self.filename)
        self.ws = self.wb.active

    def __del__(self):
        '''this functions prints filename after InputFile class closed/deleted'''
        print(self.filename)


'''------------------------------------------------------------------------------------------------------------------'''
if __name__ == '__main__':
    print("start")
    print()

    file = File('file 1.xlsx')
    print(file.ws['A1'])

    print()